import os
import math
import glob
import imghdr
import tkinter
import tkinter.filedialog
import tkinter.messagebox
from tkinter import filedialog
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from PIL import Image
import configparser

# 定数設定
SHEET_TITLE = '画像一覧' # シート名の設定
IMAGE_HEADER = 'Picture'
COMMENT_HEADER = 'Comment'
IMAGE_WIDTH = 400 # 画像の幅
IMAGE_HEIGHT = 350 # 画像の高さ
IMAGE_CELL_WIDTH = IMAGE_WIDTH * 0.05
IMAGE_CELL_HEIGHT = IMAGE_HEIGHT * 0.75

"""
init処理
設定ファイル読込
"""
if os.path.exists('./config.ini'):
    config_file = configparser.ConfigParser()
    config_file.read('./config.ini', encoding='utf-8')
    config = config_file['CONFIG']
    if config.get('SHEET_TITLE') is not None: SHEET_TITLE = config.get('SHEET_TITLE')
    if config.get('IMAGE_HEADER') is not None: IMAGE_HEADER = config.get('IMAGE_HEADER')
    if config.get('COMMENT_HEADER') is not None: COMMENT_HEADER = config.get('COMMENT_HEADER')
    if config.get('IMAGE_WIDTH') is not None: IMAGE_WIDTH = int(config.get('IMAGE_WIDTH'))
    if config.get('IMAGE_HEIGHT') is not None: IMAGE_HEIGHT = int(config.get('IMAGE_HEIGHT'))
    if config.get('IMAGE_CELL_WIDTH') is not None:
        IMAGE_CELL_WIDTH = int(config.get('IMAGE_CELL_WIDTH'))
    else:
        IMAGE_CELL_WIDTH = IMAGE_WIDTH * 0.04
    if config.get('IMAGE_CELL_HEIGHT') is not None:
        IMAGE_CELL_HEIGHT = int(config.get('IMAGE_CELL_HEIGHT'))
    else:
        IMAGE_CELL_HEIGHT = IMAGE_HEIGHT * 0.75


def main():
    """
    エントリポイント
    """
    # ワークブック設定
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0] # 1番目のシートを編集対象にする
    ws.title = SHEET_TITLE # 1番目のシートに名前を設定

    # ヘッダを描画
    write_header(ws)

    # ファイル選択ダイアログの表示
    root = tkinter.Tk()
    root.withdraw()
    iDir = os.path.abspath(os.path.dirname(__file__))
    #tkinter.messagebox.showinfo('画像一覧化App','まずは画像が置いてあるフォルダを選択してください。')
    dir = tkinter.filedialog.askdirectory(
        title = "まずは画像が置いてあるフォルダを選択してください",
        initialdir = iDir
    )

    if not dir:
        return

    # 貼り付ける画像を置いておくルートディレクトリ内のディレクトリ名を再帰的に取得
    dirs = glob.glob(os.path.join(dir, '**' + os.sep), recursive=True)

    # 各ディレクトリについて操作
    for dir_name in dirs:
        f_names = get_file_names(dir_name) # ファイル名取得
        attach_img(f_names, ws) # 画像貼り付け設定

    # ファイルへの書き込み
    filename = filedialog.asksaveasfilename(
        title = "画像一覧Excelに名前を付けて保存します",
        filetypes = [('Excelファイル', '*.xlsx')], # ファイルフィルタ
        initialdir = "./",
        initialfile = 'result.xlsx',
        defaultextension = 'xlsx'
    )
    if not filename:
        return
        
    wb.save(filename)
    tkinter.messagebox.showinfo('画像一覧化App','出力完了しました。')


def get_file_names(set_dir_name):
    """
    ディレクトリ内のファイル名取得（ファイル名のみの一覧を取得）
    """
    file_names = os.listdir(set_dir_name)
    temp_full_file_names = [os.path.join(set_dir_name, file_name) for file_name in file_names if os.path.isfile(os.path.join(set_dir_name, file_name))] # ファイルかどうかを判定
    return temp_full_file_names


def write_header(ws):
    """
    シートのヘッダ部を設定
    """
    border = Border(top=Side(style='thin', color='000000'), 
                bottom=Side(style='thin', color='000000'), 
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
    )

    ws['A1'] = SHEET_TITLE
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=16)
    
    ws['A3'] = IMAGE_HEADER
    ws['A3'].font = openpyxl.styles.Font(bold=True)
    ws['A3'].border = border
    ws['A3'].alignment = Alignment(horizontal = 'center', vertical = 'center')
    ws.merge_cells('A3:E3')
    ws['F3'] = COMMENT_HEADER
    ws['F3'].font = openpyxl.styles.Font(bold=True)
    ws['F3'].border = border
    ws['F3'].alignment = Alignment(horizontal = 'center', vertical = 'center')
    ws.merge_cells('F3:J3')


def attach_img(target_full_file_names, ws):
    """
    画像を呼び出して、Excelに貼り付け
    """
    border = Border(top=Side(style='thin', color='000000'), 
                bottom=Side(style='thin', color='000000'), 
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000')
    )
    set_row_idx = 5
    target_full_file_names.sort() # ファイル名でソート
    for target_file in target_full_file_names:
        if imghdr.what(target_file) != None: # 画像ファイルかどうかの判定
            img = openpyxl.drawing.image.Image(target_file)
            ws['A'+ str(set_row_idx - 1)] = os.path.splitext(os.path.basename(target_file))[0] # パスからファイル名のみ抽出（拡張子なし）
            ws.merge_cells('A' + str(set_row_idx - 1) + ':E' + str(set_row_idx - 1)) # 画像ファイル名のセル
            for rows in ws['A' + str(set_row_idx - 1) : 'E' + str(set_row_idx - 1)]:
                for cell in rows:
                    cell.border = border
            ws.merge_cells('A' + str(set_row_idx) + ':E' + str(set_row_idx + 20 - 1)) # 画像貼付位置のセル
            for rows in ws['A' + str(set_row_idx) : 'E' + str(set_row_idx + 20 - 1)]:
                for cell in rows:
                    cell.border = border
            ws.column_dimensions['A'].width = IMAGE_CELL_WIDTH
            ws.merge_cells('F' + str(set_row_idx - 1) + ':J' + str(set_row_idx + 20 - 1)) # コメント欄のセル
            ws['F' + str(set_row_idx - 1)].alignment = Alignment(vertical = 'top')
            for rows in ws['F' + str(set_row_idx - 1) : 'J' + str(set_row_idx + 20 - 1)]:
                for cell in rows:
                    cell.border = border
            for i in  range(20):
                ws.row_dimensions[set_row_idx + i].height = math.ceil(IMAGE_CELL_HEIGHT / 20)
            img.width = IMAGE_WIDTH
            img.height = IMAGE_HEIGHT
            ws.add_image(img, 'A' + str(set_row_idx)) # シートに画像貼り付け
            set_row_idx += 21


# 実行
main()